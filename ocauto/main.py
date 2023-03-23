import owncloud
import pyodbc
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Color
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

import os
import re
import datetime
import time
import shutil
from pathlib import Path

from utils.logger import logger
from exceptions.exceptions import ScriptNotFoundError
from exceptions.exceptions import InvalidFileExtensionError
from exceptions.exceptions import TooManyParametersError
from config import dirs
from config import database
from config import MissingConfigError

SCRIPTS_DIR = dirs.get("SCRIPTS_DIR", None)
LOCAL_ROOT_DIR = dirs.get("LOCAL_ROOT_DIR", None)
MAPPED_OC_DIRS = dirs.get("OC_MAPPED_DIRS", None)

def create_dirs_based_on_config(**kwargs):
    kwargs = dirs

    logger.info("Creating directories...")
    # Create scripts\ folder
    scripts_dir = kwargs.get("SCRIPTS_DIR", None)
    if scripts_dir is not None:
        if isinstance(scripts_dir, Path):
            scripts_dir.mkdir(exist_ok=True)
        else:
            raise TypeError(
                "'%s' must be of type 'Path', not '%s'." % ("LOCAL_ROOT_DIR", type(scripts_dir))
            )
    else:
        raise MissingConfigError(
            "'%s' is missing from your configuration file." % "SCRIPTS_DIR"
        )

    # Create local oc root dir
    local_root_dir = kwargs.get("LOCAL_ROOT_DIR", None)
    if local_root_dir is not None:
        if isinstance(local_root_dir, Path):
            local_root_dir.mkdir(exist_ok=True)
        else:
            raise TypeError(
                "'%s' must be of type 'Path', not '%s'." % ("LOCAL_ROOT_DIR", type(local_root_dir))
            )
    else:
        raise MissingConfigError(
            "'%s' is missing from your configuration file." % "LOCAL_ROOT_DIR"
        )

    oc_sub_dirs = kwargs.get("OC_MAPPED_DIRS", None)
    if oc_sub_dirs is not None:
        for key, val in oc_sub_dirs.items():
            if isinstance(val, Path):
                path = Path(LOCAL_ROOT_DIR / val)
                if not path.exists():
                    os.makedirs(path)
            else:
                raise TypeError(
                    "'%s' must be of type 'Path', not '%s'." % (key, type(val))
                )
    else:
        raise MissingConfigError(
            "'%s' is missing from your configuration file." % "OC_SUBDIRS"
        )
    logger.info("Done...")

def get_data_from_sql(script: str, param: str = None) -> pyodbc.Cursor:
    """
    This function generates a structured data (list) from a given
    sql script. This uses the python module 'pyodbc' to connect to your
    database, and perform querying tasks.
    """

    # Builds a pyodbc connection object
    logger.info("Connecting to '{database}', under instance name '{servername}'..."
        .format(database="PRD",
            servername=database.get("SERVER")
        )
    )
    conn = pyodbc.connect(
        "DRIVER={%s}; SERVER=%s; DATABASE=%s; UID=%s; PWD=%s;" 
            % (database.get("DRIVER"), database.get("SERVER"), 
            database.get("DATABASE"), database.get("UID"), database.get("PWD")),
        autocommit=False
    )

    # Creates a cursor
    cursor = conn.cursor()

    # This checks whether the passed sql script is parameterized or not
    if param is not None:
        rows = cursor.execute(script, param)
    else:
        rows = cursor.execute(script)

    data = _row_to_list(rows, include_headers=True)

    return data

def to_spreadsheet(
        data: list[list|tuple]|tuple[list|tuple], 
        local_path: str, filename: str, extension: str = "xlsx", apply_styling: bool = False,
        freeze_col_header: bool = False, auto_filter: bool = False) -> str:
    """
    Creates a spreadsheet based on the given list of data.

    params:
        data: list - the data to be converted into a spreadsheet.
        local_path: str - the path where the generated spreadsheet will be stored at.
        filename: str - the filename of your spreadsheet.
        extension: str - the file extension to be used (if isn't specified, defaults to xlsx).
        apply_styling: bool - whether to apply styling on the created spreadsheet or not.
        freeze_col_header: bool - whether to apply freeze on the created spreadsheet or not.
        auto_filter: bool - whether to apply filter on the created spreadsheet or not.

    returns:
        file: str - a string representation of the spreadsheet file location.
    """

    # Verifies if the provided file extension is a valid spreadsheet extention.
    logger.info("Validating provided file extension...")
    allowed_extensions = ["xls", "xlsx"]
    if extension not in allowed_extensions:
        raise InvalidFileExtensionError("Valid file extensions are (xls, xlsx), used '{}' instead.".format(extension))

    logger.info("Creating new workbook...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"

    for row in data:
        ws.append(row)
    
    if apply_styling:
    # Apply custom cell formatting/styling 
        logger.info("Applying styles...")
        for row in ws.iter_rows(max_row=1):
            for cell in row:
                cell.font = Font(
                    size=11,
                    bold=True
                )
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

        # This produces the alternating rows background color/shade effect for
        # each row in the spreadsheet
        rowcount = 1
        for row in ws.iter_rows(min_row=2):
            rowcount += 1
            if rowcount % 2 == 0:
                for cell in row:
                    cell.fill = PatternFill(
                        fill_type="solid",
                        fgColor="F0F0F0"
                    )
            else:
                for cell in row:
                    cell.fill = PatternFill(
                        fill_type="solid",
                        fgColor="00FFFFFF"
                    )  
        logger.info("Done...")

    # Freeze column headers
    if freeze_col_header:
        ws.freeze_panes = ws["A2"]

    # Apply auto filtering
    if auto_filter:
        ws.auto_filter.ref = ws.dimensions

    # Checks whether the given local_path, where the spreadsheet will be saved, exists. 
    # If not, creates the path accordingly.
    if not os.path.isdir(local_path):
        try:
            os.makedirs(local_path)
        except Exception as exc:
            raise exc

    # Normalize the extension
    if not extension.startswith("."):
        extension = "." + extension

    filename = local_path / (filename + extension)

    try:
        logger.info("Saving file '{}' to '{}'..."
            .format(os.path.basename(filename),
                os.path.dirname(filename)
            )
        )
        wb.save(filename)
        logger.info("Done...")
    except Exception as exc:
        raise exc

    file = filename

    return file

def copy_local_file_to_oc(source_file_path: str, destination_file_path: str):
    try:
        shutil.copy(source_file_path, destination_file_path)
    except Exception as exc:
        raise exc

def _row_to_list(row_obj: pyodbc.Cursor, include_headers: bool = False) -> list:
    """
    Transforms pyodbc.Cursor object to list.

    params:
        row_obj: pyodbc.Cursor

    returns:
        list
    """

    # Get column names
    if include_headers:
        logger.info("Extracting table columns...")
        col_names = [c[0] for c in row_obj.description]
        logger.info("Done...")

    # Get row data
    logger.info("Extracting table rows...")
    rows = [list(row) for row in row_obj]
    logger.info("Done...")

    # Merge column name(s) with row data
    data = []
    logger.info("Sanitizing report data...")
    if include_headers:
        data = list([col_names])
        for row in rows:
            data.append(row)
    else:
        for row in rows:
            data.append(row)
    logger.info("Done...")

    return data

def _get_year_and_month_before() -> str:
    """
    Transforms the date into 'YYYYMM' format
    to conform the date format to query filtering
    
    params: 
        None

    returns: 
        Formatted Date (YYYYMM)
    """

    _DATE_FORMAT = "%Y%m"

    f_date = None
    date = datetime.date.today()
    m = date.month - 1
    if isinstance(m, int) and m > 0 and m < 12: 
        f_date = date.replace(month=m)
    elif m < 1:
        m = 12
        y = date.year - 1
        f_date = date.replace(year=y, month=m)
    res = datetime.datetime.strftime(f_date, _DATE_FORMAT)
    return res

from typing import Any

def _read_file(file_path: Any) -> str:
    try:
        file_handler = open(file_path, "r")
        content = file_handler.read()
    except Exception as exc:
        raise exc
    finally:
        file_handler.close()

    return content

def _is_param_overloaded():
    pass

def _normalize_path(path: str):
    if not path.endswith("\\"):
        path += "\\"
    return path

def main():
    scripts_path = Path(SCRIPTS_DIR)

    for path in scripts_path.iterdir():
        if path.is_dir():
            raise Exception("Scripts folder must contain script files only.")

    script_files = [str(os.path.basename(file)).removesuffix(".sql").lower() for file in Path(SCRIPTS_DIR).glob("*.sql")]

    for k in MAPPED_OC_DIRS:
        if k.lower() not in script_files:
            raise ScriptNotFoundError("Script file not found for '%s'." % (k))
        try:
            script = _read_file(_normalize_path(str(SCRIPTS_DIR)) + k + ".sql")
            if re.search("= +\?", script):
                param_count = len(re.findall("= +\?", script))
                if param_count > 1:
                    raise TooManyParametersError("The function only supports %s parameter at a time, but '%s' has %s."
                        % (
                            1,
                            os.path.join(SCRIPTS_DIR, k + ".sql"),
                            param_count
                        )
                    )
                data = get_data_from_sql(script, _get_year_and_month_before())
            else:
                data = get_data_from_sql(script)

            spreadsheet = to_spreadsheet(
                data=data, 
                local_path=LOCAL_ROOT_DIR / MAPPED_OC_DIRS[k], # Points to local directory
                filename=k.upper(),
                extension="xlsx"
            )
            logger.info("Copying '{}' from '{}' to '{}'..."
                .format(os.path.basename(spreadsheet),
                    os.path.dirname(spreadsheet),
                    os.path.dirname(dirs.get("OC_ROOT_DIR") / MAPPED_OC_DIRS[k])
                )
            )
            # copy_local_file_to_oc(
            #     destination_file_path=oc_dirs_mapping["root"] + oc_dirs_mapping[k], 
            #     source_file_path=spreadsheet
            # )
            logger.info("Done...")
        except Exception as exc:
            raise exc

if __name__ == "__main__":

    # LOCAL_ROOT_DIR = dirs.get("LOCAL_ROOT", None)
    # OC_MAPPED_DIRS = dirs.get("OC_MAPPED_DIRS", None)

    # create_dirs_based_on_config(root_dir=LOCAL_ROOT_DIR, sub_dirs=OC_MAPPED_DIRS)
    create_dirs_based_on_config()
    main()